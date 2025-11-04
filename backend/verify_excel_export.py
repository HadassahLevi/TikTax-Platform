"""
Quick verification script for Excel export service
Run this to verify the implementation works
"""

import sys
import os

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from datetime import datetime, timedelta
from app.services.excel_service import excel_service
from app.models.user import User, SubscriptionPlan
from app.models.receipt import Receipt, ReceiptStatus
from app.models.category import Category
from openpyxl import load_workbook
import io

def test_excel_generation():
    """Test Excel generation without database"""
    
    print("🧪 Testing Excel Export Service...")
    
    # Create mock data
    user = User(
        id=1,
        email="test@tiktax.co.il",
        full_name="דוד כהן",
        business_name="עסק הדוגמה בע\"מ",
        business_number="123456789",
        business_type="חברה בע\"מ",
        subscription_plan=SubscriptionPlan.PRO
    )
    
    categories = [
        Category(id=1, name_hebrew="משרד", name_english="Office", icon="briefcase", color="#2563EB", is_default=True, sort_order=1),
        Category(id=2, name_hebrew="ציוד", name_english="Equipment", icon="laptop", color="#059669", is_default=True, sort_order=2),
    ]
    
    receipts = []
    base_date = datetime(2024, 1, 1)
    for i in range(5):
        receipt = Receipt(
            id=i + 1,
            user_id=1,
            original_filename=f"receipt_{i+1}.jpg",
            file_url=f"https://s3.amazonaws.com/receipts/receipt_{i+1}.jpg",
            file_size=1024 * 100,
            mime_type="image/jpeg",
            vendor_name=f"ספק מס' {i+1}",
            business_number=f"12345678{i}",
            receipt_number=f"RCP{i+1:04d}",
            receipt_date=base_date + timedelta(days=i * 3),
            total_amount=100.0 + (i * 50),
            vat_amount=14.53 + (i * 7.26),
            pre_vat_amount=85.47 + (i * 42.74),
            category_id=(i % 2) + 1,
            status=ReceiptStatus.APPROVED,
            confidence_score=95.5,
            notes=f"הערה {i+1}" if i % 2 == 0 else None,
            is_digitally_signed=True
        )
        receipts.append(receipt)
    
    # Generate Excel
    print("📊 Generating Excel export...")
    excel_bytes = excel_service.generate_export(
        user,
        receipts,
        categories,
        datetime(2024, 1, 1),
        datetime(2024, 12, 31)
    )
    
    print(f"✅ Excel generated: {len(excel_bytes)} bytes")
    
    # Verify Excel structure
    print("🔍 Verifying Excel structure...")
    wb = load_workbook(io.BytesIO(excel_bytes))
    
    assert len(wb.sheetnames) == 3, "Should have 3 sheets"
    assert "סיכום" in wb.sheetnames, "Should have Summary sheet"
    assert "פירוט קבלות" in wb.sheetnames, "Should have Details sheet"
    assert "פירוט לפי קטגוריה" in wb.sheetnames, "Should have Categories sheet"
    
    print("✅ All 3 sheets present")
    
    # Check RTL
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.sheet_view.rightToLeft is True, f"{sheet_name} should be RTL"
    
    print("✅ RTL enabled on all sheets")
    
    # Check summary content
    ws_summary = wb["סיכום"]
    all_text = " ".join([str(cell.value) for row in ws_summary.iter_rows() for cell in row if cell.value])
    assert "עסק הדוגמה בע\"מ" in all_text, "Should contain business name"
    assert "123456789" in all_text, "Should contain business number"
    
    print("✅ Summary sheet has correct business info")
    
    # Check details sheet
    ws_details = wb["פירוט קבלות"]
    data_rows = list(ws_details.iter_rows(min_row=2, values_only=True))
    non_empty_rows = [row for row in data_rows if any(row)]
    assert len(non_empty_rows) == 5, "Should have 5 receipt rows"
    
    print("✅ Details sheet has all receipts")
    
    # Save for manual inspection
    output_path = "test_export.xlsx"
    with open(output_path, 'wb') as f:
        f.write(excel_bytes)
    
    print(f"💾 Saved test export to: {output_path}")
    print("\n🎉 All tests passed! Excel export service is working correctly.")
    
    return True

if __name__ == "__main__":
    try:
        test_excel_generation()
    except Exception as e:
        print(f"\n❌ Test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
