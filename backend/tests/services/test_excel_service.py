"""
Unit Tests for Excel Service
Tests Excel generation with multi-sheet workbooks and Hebrew RTL support
"""

import pytest
from datetime import datetime, timedelta
from openpyxl import load_workbook
import io

from app.services.excel_service import excel_service, ExcelService
from app.models.user import User, SubscriptionPlan
from app.models.receipt import Receipt, ReceiptStatus
from app.models.category import Category


@pytest.fixture
def mock_user():
    """Create mock user with business info"""
    user = User(
        id=1,
        email="test@tiktax.co.il",
        full_name="דוד כהן",
        business_name="עסק הדוגמה בע\"מ",
        business_number="123456789",
        business_type="חברה בע\"מ",
        subscription_plan=SubscriptionPlan.PRO
    )
    return user


@pytest.fixture
def mock_categories():
    """Create mock categories"""
    return [
        Category(id=1, name_hebrew="משרד", name_english="Office", icon="briefcase", color="#2563EB", is_default=True, sort_order=1),
        Category(id=2, name_hebrew="ציוד", name_english="Equipment", icon="laptop", color="#059669", is_default=True, sort_order=2),
        Category(id=3, name_hebrew="נסיעות", name_english="Travel", icon="car", color="#F59E0B", is_default=True, sort_order=3),
    ]


@pytest.fixture
def mock_receipts():
    """Create mock receipts"""
    base_date = datetime(2024, 1, 1)
    receipts = []
    
    for i in range(10):
        receipt = Receipt(
            id=i + 1,
            user_id=1,
            original_filename=f"receipt_{i+1}.jpg",
            file_url=f"https://s3.amazonaws.com/receipts/receipt_{i+1}.jpg",
            file_size=1024 * 100,
            mime_type="image/jpeg",
            vendor_name=f"ספק מס' {i+1}",
            business_number=f"12345678{i}",
            receipt_number=f"RCP{i+1:04d}",
            receipt_date=base_date + timedelta(days=i * 3),
            total_amount=100.0 + (i * 50),
            vat_amount=14.53 + (i * 7.26),
            pre_vat_amount=85.47 + (i * 42.74),
            category_id=(i % 3) + 1,  # Distribute across 3 categories
            status=ReceiptStatus.APPROVED,
            confidence_score=95.5,
            notes=f"הערה {i+1}" if i % 2 == 0 else None,
            is_digitally_signed=True,
            created_at=datetime.utcnow()
        )
        receipts.append(receipt)
    
    return receipts


class TestExcelService:
    """Test suite for Excel generation service"""
    
    def test_excel_service_initialization(self):
        """Test service initializes correctly"""
        service = ExcelService()
        assert service is not None
        assert isinstance(service, ExcelService)
    
    def test_generate_export_returns_bytes(self, mock_user, mock_receipts, mock_categories):
        """Test that generate_export returns bytes"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        result = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        assert isinstance(result, bytes)
        assert len(result) > 0
    
    def test_excel_has_three_sheets(self, mock_user, mock_receipts, mock_categories):
        """Test Excel workbook has 3 sheets"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        # Load workbook from bytes
        wb = load_workbook(io.BytesIO(excel_bytes))
        
        assert len(wb.sheetnames) == 3
        assert "סיכום" in wb.sheetnames
        assert "פירוט קבלות" in wb.sheetnames
        assert "פירוט לפי קטגוריה" in wb.sheetnames
    
    def test_summary_sheet_has_correct_data(self, mock_user, mock_receipts, mock_categories):
        """Test summary sheet contains correct business info and totals"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["סיכום"]
        
        # Check RTL mode
        assert ws.sheet_view.rightToLeft is True
        
        # Check title
        assert "דוח קבלות" in ws['A1'].value
        
        # Check business info (row positions may vary, check general content)
        all_values = []
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            all_values.extend([str(v) for v in row if v])
        
        all_text = " ".join(all_values)
        assert "עסק הדוגמה בע\"מ" in all_text
        assert "123456789" in all_text
    
    def test_summary_sheet_calculates_totals_correctly(self, mock_user, mock_receipts, mock_categories):
        """Test summary sheet has correct totals"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["סיכום"]
        
        # Calculate expected totals
        expected_total = sum([r.total_amount or 0 for r in mock_receipts])
        expected_vat = sum([r.vat_amount or 0 for r in mock_receipts])
        expected_pre_vat = sum([r.pre_vat_amount or 0 for r in mock_receipts])
        
        # Find total values in sheet (scan rows for numeric values)
        found_total = False
        for row in ws.iter_rows(min_row=8, max_row=20, values_only=True):
            if row[1] and isinstance(row[1], (int, float)):
                if abs(row[1] - expected_total) < 0.01:
                    found_total = True
                    break
        
        assert found_total, f"Expected total {expected_total} not found in summary sheet"
    
    def test_details_sheet_has_all_receipts(self, mock_user, mock_receipts, mock_categories):
        """Test details sheet has all receipts"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["פירוט קבלות"]
        
        # Check RTL mode
        assert ws.sheet_view.rightToLeft is True
        
        # Count data rows (excluding header)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        non_empty_rows = [row for row in data_rows if any(row)]
        
        assert len(non_empty_rows) == len(mock_receipts)
    
    def test_details_sheet_has_correct_headers(self, mock_user, mock_receipts, mock_categories):
        """Test details sheet has correct column headers"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["פירוט קבלות"]
        
        headers = [cell.value for cell in ws[1]]
        
        assert "תאריך" in headers
        assert "ספק" in headers
        assert "מספר עוסק" in headers
        assert "מספר קבלה" in headers
        assert "קטגוריה" in headers
        assert "לפני מע\"מ" in headers
        assert "מע\"מ" in headers
        assert "סה\"כ" in headers
        assert "הערות" in headers
    
    def test_categories_sheet_groups_correctly(self, mock_user, mock_receipts, mock_categories):
        """Test categories sheet groups receipts by category"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["פירוט לפי קטגוריה"]
        
        # Check RTL mode
        assert ws.sheet_view.rightToLeft is True
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "קטגוריה" in headers
        assert "מספר קבלות" in headers
        assert "סכום כולל" in headers
        assert "אחוז" in headers
    
    def test_categories_sheet_calculates_percentages(self, mock_user, mock_receipts, mock_categories):
        """Test categories sheet has correct percentages"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["פירוט לפי קטגוריה"]
        
        # Find total row (should be last row with data)
        total_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "סה\"כ":
                total_row = row
                break
        
        assert total_row is not None
        # Total percentage should be 100% (or 1.0 in decimal)
        assert total_row[3] == 1.0 or abs(total_row[3] - 1.0) < 0.001
    
    def test_empty_receipts_list(self, mock_user, mock_categories):
        """Test handling empty receipts list"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            [],
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        
        # Should still create all 3 sheets
        assert len(wb.sheetnames) == 3
        
        # Details sheet should only have header
        ws_details = wb["פירוט קבלות"]
        data_rows = list(ws_details.iter_rows(min_row=2, values_only=True))
        non_empty_rows = [row for row in data_rows if any(row)]
        assert len(non_empty_rows) == 0
    
    def test_hebrew_text_encoding(self, mock_user, mock_receipts, mock_categories):
        """Test Hebrew text is properly encoded"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["סיכום"]
        
        # Check that Hebrew text is readable
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v and isinstance(v, str)])
        
        # Should contain Hebrew characters
        has_hebrew = any('א' <= char <= 'ת' for val in all_values for char in val)
        assert has_hebrew
    
    def test_number_formatting(self, mock_user, mock_receipts, mock_categories):
        """Test currency amounts have correct number format"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["פירוט קבלות"]
        
        # Check amount columns have currency format
        # Column F (6), G (7), H (8) are amount columns
        for row_num in range(2, min(5, len(mock_receipts) + 2)):
            for col in [6, 7, 8]:
                cell = ws.cell(row=row_num, column=col)
                # Should have currency number format
                assert '₪' in cell.number_format or '#,##0.00' in cell.number_format
    
    def test_freeze_panes_on_detail_sheets(self, mock_user, mock_receipts, mock_categories):
        """Test that detail sheets have frozen header rows"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        
        # Details sheet should have frozen first row
        ws_details = wb["פירוט קבלות"]
        assert ws_details.freeze_panes == 'A2'
        
        # Categories sheet should have frozen first row
        ws_categories = wb["פירוט לפי קטגוריה"]
        assert ws_categories.freeze_panes == 'A2'
    
    def test_column_widths_are_set(self, mock_user, mock_receipts, mock_categories):
        """Test that column widths are properly set for readability"""
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["פירוט קבלות"]
        
        # Check that columns have custom widths (not default)
        assert ws.column_dimensions['A'].width != 8.43  # Default width
        assert ws.column_dimensions['B'].width > 20  # Vendor name should be wide
        assert ws.column_dimensions['I'].width > 25  # Notes should be wide


class TestExcelServiceEdgeCases:
    """Test edge cases and error handling"""
    
    def test_receipts_with_null_amounts(self, mock_user, mock_categories):
        """Test handling receipts with null amounts"""
        receipts = [
            Receipt(
                id=1,
                user_id=1,
                original_filename="receipt.jpg",
                file_url="https://s3.amazonaws.com/receipts/receipt.jpg",
                file_size=1024,
                mime_type="image/jpeg",
                vendor_name="ספק טסט",
                receipt_date=datetime(2024, 1, 1),
                total_amount=None,
                vat_amount=None,
                pre_vat_amount=None,
                category_id=1,
                status=ReceiptStatus.APPROVED
            )
        ]
        
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        # Should not raise exception
        excel_bytes = excel_service.generate_export(
            mock_user,
            receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        assert len(excel_bytes) > 0
    
    def test_receipts_without_categories(self, mock_user, mock_categories):
        """Test handling receipts without assigned categories"""
        receipts = [
            Receipt(
                id=1,
                user_id=1,
                original_filename="receipt.jpg",
                file_url="https://s3.amazonaws.com/receipts/receipt.jpg",
                file_size=1024,
                mime_type="image/jpeg",
                vendor_name="ספק טסט",
                receipt_date=datetime(2024, 1, 1),
                total_amount=100.0,
                vat_amount=14.53,
                pre_vat_amount=85.47,
                category_id=None,  # No category
                status=ReceiptStatus.APPROVED
            )
        ]
        
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            mock_user,
            receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["פירוט קבלות"]
        
        # Category cell should show "לא מסווג"
        category_value = ws.cell(row=2, column=5).value
        assert category_value == "לא מסווג" or category_value == ""
    
    def test_user_without_business_info(self, mock_categories, mock_receipts):
        """Test handling user without business information"""
        user = User(
            id=1,
            email="test@tiktax.co.il",
            full_name="משתמש טסט",
            business_name=None,
            business_number=None,
            business_type=None
        )
        
        date_from = datetime(2024, 1, 1)
        date_to = datetime(2024, 12, 31)
        
        excel_bytes = excel_service.generate_export(
            user,
            mock_receipts,
            mock_categories,
            date_from,
            date_to
        )
        
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["סיכום"]
        
        # Should show "לא צוין" for missing fields
        all_values = [str(cell.value) for row in ws.iter_rows(values_only=True) for cell in row if cell]
        assert "לא צוין" in " ".join(all_values)
