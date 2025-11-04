"""
Integration Tests for Export Endpoints
Tests the complete export flow including authentication, generation, and download
"""

import pytest
from datetime import datetime, timedelta
from fastapi import status
from openpyxl import load_workbook
import io

from app.models.receipt import ReceiptStatus
from app.models.category import Category


class TestExportGeneration:
    """Test export generation endpoint"""
    
    @pytest.fixture(autouse=True)
    def setup(self, db, test_user, auth_headers):
        """Setup test data"""
        self.db = db
        self.user = test_user
        self.headers = auth_headers
        
        # Create categories
        categories = [
            Category(id=1, name_hebrew="משרד", name_english="Office", icon="briefcase", color="#2563EB"),
            Category(id=2, name_hebrew="ציוד", name_english="Equipment", icon="laptop", color="#059669"),
            Category(id=3, name_hebrew="נסיעות", name_english="Travel", icon="car", color="#F59E0B"),
        ]
        for cat in categories:
            db.add(cat)
        db.commit()
    
    def test_generate_excel_export_success(self, client, test_receipts):
        """Test successful Excel export generation"""
        # Ensure receipts are APPROVED
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59",
            "category_ids": None,
            "include_images": False
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_201_CREATED
        data = response.json()
        
        assert "export_id" in data
        assert "download_url" in data
        assert "expires_at" in data
        assert "file_size" in data
        assert data["file_size"] > 0
        assert "message" in data
        assert "הקובץ הופק בהצלחה" in data["message"]
    
    def test_generate_csv_export_success(self, client, test_receipts):
        """Test successful CSV export generation"""
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "csv",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_201_CREATED
        data = response.json()
        
        assert data["file_size"] > 0
        assert ".csv" in data["download_url"]
    
    def test_generate_export_with_category_filter(self, client, test_receipts):
        """Test export with category filter"""
        for receipt in test_receipts[:5]:
            receipt.status = ReceiptStatus.APPROVED
            receipt.category_id = 1
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59",
            "category_ids": [1]
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_201_CREATED
    
    def test_generate_export_invalid_date_range(self, client):
        """Test export with invalid date range (start after end)"""
        payload = {
            "format": "excel",
            "date_from": "2024-12-31T00:00:00",
            "date_to": "2024-01-01T00:00:00"  # End before start
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "תאריך התחלה" in response.json()["detail"]
    
    def test_generate_export_date_range_too_large(self, client):
        """Test export with date range > 2 years"""
        payload = {
            "format": "excel",
            "date_from": "2020-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"  # ~5 years
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "טווח תאריכים" in response.json()["detail"]
    
    def test_generate_export_no_receipts_found(self, client):
        """Test export when no receipts match criteria"""
        payload = {
            "format": "excel",
            "date_from": "2025-01-01T00:00:00",
            "date_to": "2025-12-31T23:59:59"  # Future dates
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_404_NOT_FOUND
        assert "לא נמצאו קבלות" in response.json()["detail"]
    
    def test_generate_export_only_approved_receipts(self, client, test_receipts):
        """Test export only includes APPROVED receipts"""
        # Set different statuses
        test_receipts[0].status = ReceiptStatus.APPROVED
        test_receipts[1].status = ReceiptStatus.REVIEW
        test_receipts[2].status = ReceiptStatus.PROCESSING
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_201_CREATED
        # Should only include 1 receipt (APPROVED)
        assert "1 קבלות" in response.json()["message"]
    
    def test_generate_export_pdf_not_implemented(self, client, test_receipts):
        """Test PDF export returns not implemented"""
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "pdf",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        assert response.status_code == status.HTTP_501_NOT_IMPLEMENTED
        assert "PDF" in response.json()["detail"]
    
    def test_generate_export_requires_authentication(self, client):
        """Test export requires authentication"""
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        response = client.post("/api/v1/export/generate", json=payload)
        
        assert response.status_code == status.HTTP_401_UNAUTHORIZED


class TestExportDownload:
    """Test export download endpoint"""
    
    @pytest.fixture(autouse=True)
    def setup(self, db, test_user, auth_headers):
        """Setup test data"""
        self.db = db
        self.user = test_user
        self.headers = auth_headers
        
        # Create categories
        categories = [
            Category(id=1, name_hebrew="משרד", name_english="Office", icon="briefcase", color="#2563EB"),
        ]
        for cat in categories:
            db.add(cat)
        db.commit()
    
    def test_download_export_success(self, client, test_receipts):
        """Test successful export download"""
        # Generate export first
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        gen_response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        assert gen_response.status_code == status.HTTP_201_CREATED
        
        export_data = gen_response.json()
        download_url = export_data["download_url"]
        
        # Download the file
        response = client.get(download_url, headers=self.headers)
        
        assert response.status_code == status.HTTP_200_OK
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]
        
        # Verify it's a valid Excel file
        excel_bytes = response.content
        wb = load_workbook(io.BytesIO(excel_bytes))
        assert len(wb.sheetnames) == 3
    
    def test_download_export_csv_success(self, client, test_receipts):
        """Test successful CSV download"""
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "csv",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        gen_response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        download_url = gen_response.json()["download_url"]
        
        response = client.get(download_url, headers=self.headers)
        
        assert response.status_code == status.HTTP_200_OK
        assert "text/csv" in response.headers["content-type"]
        assert ".csv" in response.headers["content-disposition"]
        
        # Verify CSV content
        csv_content = response.content.decode('utf-8-sig')
        assert "תאריך" in csv_content
        assert "ספק" in csv_content
    
    def test_download_export_not_found(self, client):
        """Test download with invalid export ID"""
        response = client.get("/api/v1/export/download/invalid-uuid", headers=self.headers)
        
        assert response.status_code == status.HTTP_404_NOT_FOUND
        assert "לא נמצא" in response.json()["detail"]
    
    def test_download_export_unauthorized_access(self, client, test_receipts, test_user_2, auth_headers_2):
        """Test user cannot download another user's export"""
        # User 1 generates export
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        gen_response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        download_url = gen_response.json()["download_url"]
        
        # User 2 tries to download
        response = client.get(download_url, headers=auth_headers_2)
        
        assert response.status_code == status.HTTP_403_FORBIDDEN
        assert "אין הרשאה" in response.json()["detail"]
    
    def test_download_export_requires_authentication(self, client):
        """Test download requires authentication"""
        response = client.get("/api/v1/export/download/some-uuid")
        
        assert response.status_code == status.HTTP_401_UNAUTHORIZED
    
    def test_download_export_cache_headers(self, client, test_receipts):
        """Test download has proper cache-control headers"""
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        gen_response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        download_url = gen_response.json()["download_url"]
        
        response = client.get(download_url, headers=self.headers)
        
        assert response.status_code == status.HTTP_200_OK
        assert "no-cache" in response.headers["cache-control"]
        assert "no-store" in response.headers["cache-control"]


class TestExportCleanup:
    """Test export cleanup endpoint"""
    
    @pytest.fixture(autouse=True)
    def setup(self, db, test_user, auth_headers):
        """Setup test data"""
        self.db = db
        self.user = test_user
        self.headers = auth_headers
        
        # Create categories
        categories = [
            Category(id=1, name_hebrew="משרד", name_english="Office", icon="briefcase", color="#2563EB"),
        ]
        for cat in categories:
            db.add(cat)
        db.commit()
    
    def test_cleanup_expired_exports(self, client, test_receipts):
        """Test cleanup removes expired exports"""
        # Generate multiple exports
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        # Generate 3 exports
        for _ in range(3):
            client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        
        # Cleanup
        response = client.delete("/api/v1/export/cleanup", headers=self.headers)
        
        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert "cleaned_count" in data
        assert "remaining_count" in data
        assert "message" in data


class TestExportContentValidation:
    """Test the actual content of generated exports"""
    
    @pytest.fixture(autouse=True)
    def setup(self, db, test_user, auth_headers):
        """Setup test data"""
        self.db = db
        self.user = test_user
        self.headers = auth_headers
        
        # Create categories
        categories = [
            Category(id=1, name_hebrew="משרד", name_english="Office", icon="briefcase", color="#2563EB"),
            Category(id=2, name_hebrew="ציוד", name_english="Equipment", icon="laptop", color="#059669"),
        ]
        for cat in categories:
            db.add(cat)
        db.commit()
    
    def test_excel_contains_business_info(self, client, test_receipts):
        """Test Excel export contains user's business information"""
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        # Update user business info
        self.user.business_name = "חברת הבדיקה בע\"מ"
        self.user.business_number = "987654321"
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        gen_response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        download_url = gen_response.json()["download_url"]
        
        response = client.get(download_url, headers=self.headers)
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["סיכום"]
        
        # Check business info is present
        all_text = " ".join([str(cell.value) for row in ws.iter_rows() for cell in row if cell.value])
        assert "חברת הבדיקה בע\"מ" in all_text
        assert "987654321" in all_text
    
    def test_excel_hebrew_rtl_enabled(self, client, test_receipts):
        """Test all sheets have RTL enabled"""
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "excel",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        gen_response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        download_url = gen_response.json()["download_url"]
        
        response = client.get(download_url, headers=self.headers)
        wb = load_workbook(io.BytesIO(response.content))
        
        # All sheets should have RTL
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert ws.sheet_view.rightToLeft is True
    
    def test_csv_has_hebrew_bom(self, client, test_receipts):
        """Test CSV has BOM for proper Excel Hebrew display"""
        for receipt in test_receipts:
            receipt.status = ReceiptStatus.APPROVED
        self.db.commit()
        
        payload = {
            "format": "csv",
            "date_from": "2024-01-01T00:00:00",
            "date_to": "2024-12-31T23:59:59"
        }
        
        gen_response = client.post("/api/v1/export/generate", json=payload, headers=self.headers)
        download_url = gen_response.json()["download_url"]
        
        response = client.get(download_url, headers=self.headers)
        csv_bytes = response.content
        
        # Should start with UTF-8 BOM
        assert csv_bytes[:3] == b'\xef\xbb\xbf'
