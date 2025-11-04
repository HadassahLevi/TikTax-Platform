"""
Excel Export Service
Generate multi-sheet Excel workbooks with receipt data for accountants
Includes Hebrew RTL support, professional formatting, and Israeli tax compliance
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List
import io
import logging

from ..models.receipt import Receipt
from ..models.category import Category
from ..models.user import User
from ..utils.formatters import format_amount, format_israeli_date

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for generating professional Excel exports"""
    
    def generate_export(
        self,
        user: User,
        receipts: List[Receipt],
        categories: List[Category],
        date_from: datetime,
        date_to: datetime
    ) -> bytes:
        """
        Generate Excel workbook with 3 sheets:
        1. Summary - Business info and totals
        2. Details - All receipts
        3. Categories - Breakdown by category
        
        Args:
            user: User generating the export
            receipts: List of receipts to include
            categories: All categories for lookup
            date_from: Report start date
            date_to: Report end date
            
        Returns:
            Excel file as bytes
        """
        logger.info(f"Generating Excel export for user {user.id} with {len(receipts)} receipts")
        
        wb = Workbook()
        
        # Create sheets (in reverse order, Summary will be first)
        self._create_categories_sheet(wb, receipts, categories)
        self._create_details_sheet(wb, receipts, categories)
        self._create_summary_sheet(wb, user, receipts, date_from, date_to)
        
        # Remove default sheet if it exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel export generated successfully, size: {len(output.getvalue())} bytes")
        
        return output.read()
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        user: User,
        receipts: List[Receipt],
        date_from: datetime,
        date_to: datetime
    ):
        """Sheet 1: Summary with business info and totals"""
        # Create as first sheet
        ws = wb.create_sheet("סיכום", 0)
        ws.sheet_view.rightToLeft = True  # RTL for Hebrew
        
        # Header styling
        header_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        
        # Title
        ws['A1'] = "דוח קבלות - Tik-Tax"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Business info section
        row = 3
        info_label_font = Font(name='Arial', size=11, bold=True)
        info_value_font = Font(name='Arial', size=11)
        
        ws[f'A{row}'] = "שם העסק:"
        ws[f'B{row}'] = user.business_name or "לא צוין"
        ws[f'A{row}'].font = info_label_font
        ws[f'B{row}'].font = info_value_font
        
        row += 1
        ws[f'A{row}'] = "מספר עוסק:"
        ws[f'B{row}'] = user.business_number or "לא צוין"
        ws[f'A{row}'].font = info_label_font
        ws[f'B{row}'].font = info_value_font
        
        row += 1
        ws[f'A{row}'] = "סוג עסק:"
        ws[f'B{row}'] = user.business_type or "לא צוין"
        ws[f'A{row}'].font = info_label_font
        ws[f'B{row}'].font = info_value_font
        
        row += 1
        ws[f'A{row}'] = "תקופת הדוח:"
        ws[f'B{row}'] = f"{format_israeli_date(date_from)} - {format_israeli_date(date_to)}"
        ws[f'A{row}'].font = info_label_font
        ws[f'B{row}'].font = info_value_font
        
        row += 1
        ws[f'A{row}'] = "תאריך יצירה:"
        ws[f'B{row}'] = format_israeli_date(datetime.utcnow())
        ws[f'A{row}'].font = info_label_font
        ws[f'B{row}'].font = info_value_font
        
        # Totals section
        row += 2
        totals_header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        totals_header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        
        ws[f'A{row}'] = "סיכום כספי"
        ws[f'A{row}'].font = totals_header_font
        ws[f'A{row}'].fill = totals_header_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 25
        
        row += 1
        
        # Calculate totals
        total_amount = sum([r.total_amount or 0 for r in receipts])
        total_vat = sum([r.vat_amount or 0 for r in receipts])
        total_pre_vat = sum([r.pre_vat_amount or 0 for r in receipts])
        
        totals_label_font = Font(name='Arial', size=12, bold=True)
        totals_value_font = Font(name='Arial', size=12)
        
        ws[f'A{row}'] = "סה\"כ קבלות:"
        ws[f'B{row}'] = len(receipts)
        ws[f'A{row}'].font = totals_label_font
        ws[f'B{row}'].font = totals_value_font
        
        row += 1
        ws[f'A{row}'] = "סה\"כ לפני מע\"מ:"
        ws[f'B{row}'] = total_pre_vat
        ws[f'A{row}'].font = totals_label_font
        ws[f'B{row}'].font = totals_value_font
        ws[f'B{row}'].number_format = '₪#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "סה\"כ מע\"מ:"
        ws[f'B{row}'] = total_vat
        ws[f'A{row}'].font = totals_label_font
        ws[f'B{row}'].font = totals_value_font
        ws[f'B{row}'].number_format = '₪#,##0.00'
        
        row += 1
        # Highlight total row
        grand_total_font = Font(name='Arial', size=14, bold=True)
        grand_total_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        
        ws[f'A{row}'] = "סה\"כ כולל מע\"מ:"
        ws[f'B{row}'] = total_amount
        ws[f'A{row}'].font = grand_total_font
        ws[f'B{row}'].font = grand_total_font
        ws[f'B{row}'].number_format = '₪#,##0.00'
        ws[f'A{row}'].fill = grand_total_fill
        ws[f'B{row}'].fill = grand_total_fill
        
        # Footer note
        row += 3
        footer_font = Font(name='Arial', size=10, italic=True, color="6B7280")
        ws[f'A{row}'] = "דוח זה הופק באמצעות Tik-Tax - מערכת ניהול קבלות חכמה"
        ws[f'A{row}'].font = footer_font
        ws.merge_cells(f'A{row}:D{row}')
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_details_sheet(self, wb: Workbook, receipts: List[Receipt], categories: List[Category]):
        """Sheet 2: Detailed receipts list"""
        ws = wb.create_sheet("פירוט קבלות")
        ws.sheet_view.rightToLeft = True
        
        # Headers
        headers = [
            "תאריך", "ספק", "מספר עוסק", "מספר קבלה",
            "קטגוריה", "לפני מע\"מ", "מע\"מ", "סה\"כ", "הערות"
        ]
        
        # Header styling
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color="D1D5DB"),
            right=Side(style='thin', color="D1D5DB"),
            top=Side(style='thin', color="D1D5DB"),
            bottom=Side(style='thin', color="D1D5DB")
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        # Category lookup
        category_dict = {cat.id: cat.name_hebrew for cat in categories}
        
        # Data rows
        data_font = Font(name='Arial', size=10)
        for row_num, receipt in enumerate(receipts, 2):
            ws.cell(row=row_num, column=1, value=format_israeli_date(receipt.receipt_date) if receipt.receipt_date else "")
            ws.cell(row=row_num, column=2, value=receipt.vendor_name or "")
            ws.cell(row=row_num, column=3, value=receipt.business_number or "")
            ws.cell(row=row_num, column=4, value=receipt.receipt_number or "")
            ws.cell(row=row_num, column=5, value=category_dict.get(receipt.category_id, "לא מסווג"))
            ws.cell(row=row_num, column=6, value=receipt.pre_vat_amount or 0)
            ws.cell(row=row_num, column=7, value=receipt.vat_amount or 0)
            ws.cell(row=row_num, column=8, value=receipt.total_amount or 0)
            ws.cell(row=row_num, column=9, value=receipt.notes or "")
            
            # Apply font and alignment
            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col >= 6 and col <= 8:  # Amount columns
                    cell.alignment = Alignment(horizontal='right')
            
            # Number formatting for amounts
            for col in [6, 7, 8]:
                ws.cell(row=row_num, column=col).number_format = '₪#,##0.00'
        
        # Column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 25  # Vendor
        ws.column_dimensions['C'].width = 12  # Business #
        ws.column_dimensions['D'].width = 12  # Receipt #
        ws.column_dimensions['E'].width = 15  # Category
        ws.column_dimensions['F'].width = 14  # Pre-VAT
        ws.column_dimensions['G'].width = 12  # VAT
        ws.column_dimensions['H'].width = 14  # Total
        ws.column_dimensions['I'].width = 30  # Notes
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_categories_sheet(self, wb: Workbook, receipts: List[Receipt], categories: List[Category]):
        """Sheet 3: Category breakdown"""
        ws = wb.create_sheet("פירוט לפי קטגוריה")
        ws.sheet_view.rightToLeft = True
        
        # Group receipts by category
        from collections import defaultdict
        category_data = defaultdict(lambda: {'count': 0, 'total': 0.0})
        
        for receipt in receipts:
            if receipt.category_id and receipt.total_amount:
                category_data[receipt.category_id]['count'] += 1
                category_data[receipt.category_id]['total'] += receipt.total_amount
        
        # Headers
        headers = ["קטגוריה", "מספר קבלות", "סכום כולל", "אחוז"]
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color="D1D5DB"),
            right=Side(style='thin', color="D1D5DB"),
            top=Side(style='thin', color="D1D5DB"),
            bottom=Side(style='thin', color="D1D5DB")
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        # Calculate total
        grand_total = sum([data['total'] for data in category_data.values()])
        
        # Category lookup
        category_dict = {cat.id: cat.name_hebrew for cat in categories}
        
        # Data rows
        data_font = Font(name='Arial', size=10)
        row_num = 2
        for cat_id, data in sorted(category_data.items(), key=lambda x: x[1]['total'], reverse=True):
            ws.cell(row=row_num, column=1, value=category_dict.get(cat_id, "לא מסווג"))
            ws.cell(row=row_num, column=2, value=data['count'])
            ws.cell(row=row_num, column=3, value=data['total'])
            
            percentage = (data['total'] / grand_total * 100) if grand_total > 0 else 0
            ws.cell(row=row_num, column=4, value=percentage / 100)  # Excel percentage format
            
            # Apply styling
            for col in range(1, 5):
                cell = ws.cell(row=row_num, column=col)
                cell.font = data_font
                cell.border = thin_border
                if col == 2:
                    cell.alignment = Alignment(horizontal='center')
                elif col >= 3:
                    cell.alignment = Alignment(horizontal='right')
            
            # Number formatting
            ws.cell(row=row_num, column=3).number_format = '₪#,##0.00'
            ws.cell(row=row_num, column=4).number_format = '0.0%'
            
            row_num += 1
        
        # Total row
        total_font = Font(name='Arial', size=11, bold=True)
        total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        ws.cell(row=row_num, column=1, value="סה\"כ")
        ws.cell(row=row_num, column=2, value=len(receipts))
        ws.cell(row=row_num, column=3, value=grand_total)
        ws.cell(row=row_num, column=4, value=1.0)  # 100%
        
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            if col == 2:
                cell.alignment = Alignment(horizontal='center')
            elif col >= 3:
                cell.alignment = Alignment(horizontal='right')
        
        ws.cell(row=row_num, column=3).number_format = '₪#,##0.00'
        ws.cell(row=row_num, column=4).number_format = '0.0%'
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        
        # Freeze header row
        ws.freeze_panes = 'A2'


# Singleton instance
excel_service = ExcelService()
